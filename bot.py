from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import service
from selenium import webdriver
from openpyxl import load_workbook
import time
import os
import sys

def LerConfig():
    # Func criada para ler linha a linha do settings.xlsx e adiciona-los em um dict
    config = load_workbook(filename="settings.xlsx")
    settings = config.get_sheet_by_name("Settings")

    dicConfig = dict()

    var_intTotalRows = settings.max_row
    for i in range(1, var_intTotalRows):
        key = settings["A" + (i+1).__str__()].value
        value = settings["B" + (i+1).__str__()].value
        if key != None and value != None: 
            dicConfig[key] = value
    return dicConfig
